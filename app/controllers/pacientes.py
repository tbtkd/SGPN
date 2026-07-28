from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from werkzeug.utils import secure_filename
import os
import openpyxl
from datetime import datetime
from app import db_orm as db
from app.models.paciente import Paciente
from app.models.pago import Pago
from app.models.historial_clinico import HistorialClinico
from app.models.valoracion_antropometrica import ValoracionAntropometrica
from app.models.cita import Cita

pacientes = Blueprint('pacientes', __name__, url_prefix='/pacientes')

@pacientes.route('/nuevo', methods=['GET', 'POST'])
def nuevo_paciente():
    if request.method == 'POST':
        try:
            telefono = request.form.get('telefono', '').strip()
            # Validar teléfono
            if len(telefono) != 10 or not telefono.isdigit():
                flash('El teléfono debe tener exactamente 10 dígitos numéricos', 'error')
                return render_template('pacientes/nuevo_paciente.html')
            
            exito, mensaje = Paciente.crear(
                request.form['nombre'],
                request.form['apellido_paterno'],
                request.form['apellido_materno'],
                request.form['genero'],
                request.form['fecha_nacimiento'],
                telefono,
                request.form['correo'],
                request.form['ciudad']
            )
            
            if exito:
                flash('Paciente registrado exitosamente', 'success')
                return redirect(url_for('pacientes.lista_pacientes_activos'))
            else:
                flash(f'Error al registrar el paciente: {mensaje}', 'error')
                return render_template('pacientes/nuevo_paciente.html')
        except Exception as e:
            flash(f'Error al registrar el paciente: {str(e)}', 'error')
            return render_template('pacientes/nuevo_paciente.html')
    
    return render_template('pacientes/nuevo_paciente.html')

@pacientes.route('/activos')
def lista_pacientes_activos():
    try:
        busqueda = request.args.get('busqueda', '')
        pacientes_list = Paciente.buscar(busqueda, status='activo')
        return render_template('pacientes/lista_pacientes.html', 
                            pacientes=pacientes_list,
                            busqueda=busqueda,
                            tipo_lista="activos")
    except Exception as e:
        flash(f'Error al obtener la lista de pacientes: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@pacientes.route('/inactivos')
def lista_pacientes_inactivos():
    try:
        busqueda = request.args.get('busqueda', '')
        pacientes_list = Paciente.buscar(busqueda, status='inactivo')
        return render_template('pacientes/lista_pacientes.html', 
                            pacientes=pacientes_list,
                            busqueda=busqueda,
                            tipo_lista="inactivos")
    except Exception as e:
        flash(f'Error al obtener la lista de pacientes: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@pacientes.route('/<int:id>', methods=['GET', 'POST'])
def detalle_paciente(id):
    paciente = Paciente.obtener_por_id(id)
    if paciente is None:
        flash('Paciente no encontrado', 'error')
        return redirect(url_for('pacientes.lista_pacientes_activos'))
    
    if request.method == 'POST':
        if 'fecha_pago' in request.form:
            fecha_pago = request.form['fecha_pago']
            try:
                Pago.registrar(id, fecha_pago)
                flash('Pago registrado exitosamente', 'success')
            except Exception as e:
                flash(f'Error al registrar el pago: {str(e)}', 'error')
        elif 'ultima_dieta' in request.form:
            ultima_dieta = request.form['ultima_dieta']
            try:
                exito, mensaje = ValoracionAntropometrica.actualizar_ultima_dieta(id, ultima_dieta)
                if exito:
                    flash(mensaje, 'success')
                else:
                    flash(mensaje, 'error')
            except Exception as e:
                flash(f'Error al actualizar la última dieta: {str(e)}', 'error')
        return redirect(url_for('pacientes.detalle_paciente', id=id))

    ultimo_pago = Pago.obtener_ultimo_pago(id)
    historial = HistorialClinico.obtener_por_paciente_id(id)
    
    valoraciones = ValoracionAntropometrica.obtener_por_paciente(id)
    ultima_valoracion = valoraciones[0] if len(valoraciones) > 0 else None
    valoracion_anterior = valoraciones[1] if len(valoraciones) > 1 else None
    
    diferencias = {}
    if ultima_valoracion and valoracion_anterior:
        campos = ['cintura', 'torax', 'brazo', 'bicep', 'tricep', 'cadera', 'pierna', 'pantorrilla', 'subescapular', 'suprailiaco', 'femoral']
        for campo in campos:
            val_actual = getattr(ultima_valoracion, campo, None)
            val_anterior = getattr(valoracion_anterior, campo, None)
            if val_actual is not None and val_anterior is not None:
                diff = val_actual - val_anterior
                tendencia = 'aumento' if diff > 0 else ('reduccion' if diff < 0 else 'sin_cambio')
                diferencias[campo] = {'valor': diff, 'tendencia': tendencia}

    siguiente_cita_data = Cita.obtener_siguiente_cita(id)
    siguiente_cita = None
    if siguiente_cita_data:
        fecha_str = str(siguiente_cita_data.fecha)
        hora_str = str(siguiente_cita_data.hora)
        hora_limpia = ":".join(hora_str.split(':')[:2])
        siguiente_cita = datetime.strptime(f"{fecha_str} {hora_limpia}", '%Y-%m-%d %H:%M')

    today = datetime.now()
    return render_template('pacientes/detalle_paciente.html', 
                            paciente=paciente, 
                            ultimo_pago=ultimo_pago,
                            historial=historial,
                            ultima_valoracion=ultima_valoracion,
                            valoracion_anterior=valoracion_anterior,
                            diferencias=diferencias,
                            siguiente_cita=siguiente_cita,
                            today=today)

@pacientes.route('/<int:id>/editar', methods=['GET', 'POST'])
def editar_paciente(id):
    if request.method == 'POST':
        try:
            telefono = request.form.get('telefono', '').strip()
            if len(telefono) != 10 or not telefono.isdigit():
                flash('El teléfono debe tener exactamente 10 dígitos numéricos', 'error')
                return redirect(url_for('pacientes.editar_paciente', id=id))
            
            status = request.form.get('status')
            if status not in ['activo', 'inactivo']:
                flash('El estado debe ser activo o inactivo', 'error')
                return redirect(url_for('pacientes.editar_paciente', id=id))
            
            Paciente.actualizar(
                id,
                request.form['nombre'],
                request.form['apellido_paterno'],
                request.form['apellido_materno'],
                request.form['genero'],
                request.form['fecha_nacimiento'],
                telefono,
                request.form['correo'],
                request.form['ciudad'],
                status
            )
            flash('Paciente actualizado exitosamente', 'success')
            return redirect(url_for('pacientes.detalle_paciente', id=id))
        except Exception as e:
            flash(f'Error al actualizar el paciente: {str(e)}', 'error')
            return redirect(url_for('pacientes.editar_paciente', id=id))
    
    paciente = Paciente.obtener_por_id(id)
    if paciente is None:
        flash('Paciente no encontrado', 'error')
        return redirect(url_for('pacientes.lista_pacientes_activos'))
    
    return render_template('pacientes/editar_paciente.html', paciente=paciente)

@pacientes.route('/<int:id>/pago', methods=['POST'])
def registrar_pago_paciente(id):
    fecha_pago = request.form['fecha_pago']
    Pago.registrar(id, fecha_pago)
    flash('Pago registrado exitosamente', 'success')
    return redirect(url_for('pacientes.detalle_paciente', id=id))

@pacientes.route('/<int:id>/cambiar-estado', methods=['POST'])
def cambiar_estado(id):
    try:
        paciente = Paciente.obtener_por_id(id)
        if not paciente:
            raise Exception('Paciente no encontrado')
        
        nuevo_estado = 'inactivo' if paciente.status == 'activo' else 'activo'
        Paciente.actualizar_estatus(id, nuevo_estado)
        
        flash('Estado del paciente actualizado correctamente', 'success')
        return jsonify({'success': True, 'nuevo_estado': nuevo_estado})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

def extraer_valor(texto, clave):
    try:
        partes = texto.lower().split(clave)
        if len(partes) > 1:
            for item in partes[1].split():
                try:
                    return float(item)
                except ValueError:
                    continue
    except Exception as e:
        print(f"Error extrayendo {clave}: {e}")
    return None

def convertir_fecha(fecha_valor):
    if isinstance(fecha_valor, datetime):
        return fecha_valor.date()
    elif isinstance(fecha_valor, str):
        try:
            partes = fecha_valor.split()
            fecha_limpia = partes[0] if len(partes) > 0 else fecha_valor
            return datetime.strptime(fecha_limpia, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None

def existe_registro(paciente_id, numero_cita, fecha):
    existente = ValoracionAntropometrica.query.filter_by(
        paciente_id=paciente_id,
        numero_cita=numero_cita,
        fecha=fecha
    ).first()
    return existente is not None

@pacientes.route('/<int:id>/cargar-excel', methods=['POST'])
def cargar_excel(id):
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
    
    file = request.files['excel_file']
    if file.filename == '' or not file.filename.endswith(('.xls', '.xlsx')):
        return jsonify({'success': False, 'message': 'El archivo debe ser un Excel (.xls o .xlsx)'})

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        estatura = sheet['M8'].value if sheet['M8'].value else None
        row = 10
        registros_procesados = 0
        registros_duplicados = 0
        errores = []
        
        while True:
            if not sheet.cell(row=row, column=12).value:
                break
            
            try:
                fecha = convertir_fecha(sheet.cell(row=row, column=13).value)
                if not fecha:
                    errores.append(f"Error en fila {row}: Formato de fecha inválido")
                    row += 1
                    continue
                
                numero_cita = sheet.cell(row=row, column=12).value
                if existe_registro(id, numero_cita, fecha):
                    registros_duplicados += 1
                    row += 1
                    continue
                
                datos = {
                    'paciente_id': id,
                    'numero_cita': numero_cita,
                    'fecha': fecha,
                    'estatura': estatura,
                    'peso': sheet.cell(row=row, column=14).value,
                    'imc': extraer_valor(str(sheet.cell(row=row, column=15).value or ''), 'imc'),
                    'grasa': extraer_valor(str(sheet.cell(row=row, column=15).value or ''), 'grasa'),
                    'cintura': sheet.cell(row=row, column=16).value,
                    'torax': sheet.cell(row=row, column=17).value,
                    'brazo': sheet.cell(row=row, column=18).value,
                    'cadera': sheet.cell(row=row, column=19).value,
                    'pierna': sheet.cell(row=row, column=20).value,
                    'pantorrilla': sheet.cell(row=row, column=21).value,
                    'tension_arterial': sheet.cell(row=row, column=24).value,
                    'frecuencia_cardiaca': sheet.cell(row=row, column=25).value,
                    'bicep': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'bc'),
                    'tricep': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'tc'),
                    'suprailiaco': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'si'),
                    'subescapular': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'se'),
                    'femoral': extraer_valor(str(sheet.cell(row=row, column=22).value or ''), 'fem'),
                    'porcentaje_grasa': sheet.cell(row=row, column=23).value,
                    'ultima_dieta': None
                }
                
                exito, mensaje = ValoracionAntropometrica.crear(id, datos)
                if exito:
                    registros_procesados += 1
                else:
                    errores.append(f"Error en fila {row}: {mensaje}")
                
            except Exception as e:
                errores.append(f"Error en fila {row}: {str(e)}")
            
            row += 1
        
        mensaje = ''
        if registros_duplicados > 0:
            mensaje = f'Se encontraron {registros_duplicados} registros que ya existen en el sistema. '
        if registros_procesados > 0:
            mensaje += f'Se agregaron {registros_procesados} nuevos registros.'
        elif registros_duplicados > 0:
            mensaje += 'No se agregaron nuevos registros.'
        
        resultado = {
            'success': True,
            'message': mensaje,
            'registros_duplicados': registros_duplicados,
            'registros_procesados': registros_procesados,
            'errores': errores if errores else "No se encontraron errores."
        }
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al procesar el archivo: {str(e)}'})

@pacientes.route('/<int:id>/registrar_proxima_cita', methods=['POST'])
def registrar_proxima_cita(id):
    fecha = request.form.get('proxima_cita_fecha', '').strip()
    hora = request.form.get('proxima_cita_hora', '').strip()

    if not fecha or not hora:
        flash('La fecha y la hora de la cita son campos obligatorios.', 'error')
        return redirect(url_for('pacientes.detalle_paciente', id=id))

    if datetime.strptime(fecha, '%Y-%m-%d').date() < datetime.now().date():
        flash('No se pueden agendar citas en fechas anteriores al día de hoy.', 'error')
        return redirect(url_for('pacientes.detalle_paciente', id=id))

    try:
        partes = hora.split(':')
        if len(partes) < 2:
            raise ValueError
        hora_int = int(partes[0])
    except (ValueError, AttributeError):
        flash('El formato de la hora proporcionada es inválido.', 'error')
        return redirect(url_for('pacientes.detalle_paciente', id=id))

    if not (9 <= hora_int <= 19):
        flash('La hora de la cita debe estar entre las 9:00 AM y las 7:00 PM.', 'error')
        return redirect(url_for('pacientes.detalle_paciente', id=id))
        
    cita_pendiente = Cita.obtener_cita_pendiente(id)
    
    if cita_pendiente:
        cita_id = cita_pendiente.id
        if not Cita.es_horario_disponible(fecha, hora, excluir_cita_id=cita_id):
            flash('El horario seleccionado ya no está disponible.', 'error')
            return redirect(url_for('pacientes.detalle_paciente', id=id))
            
        cita_pendiente.fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
        cita_pendiente.hora = datetime.strptime(hora, '%H:%M').time() if len(hora) == 5 else datetime.strptime(hora, '%H:%M:%S').time()
        db.session.commit()
        flash('Cita actualizada exitosamente.', 'success')
    else:
        if not Cita.es_horario_disponible(fecha, hora):
            flash('El horario seleccionado ya no está disponible.', 'error')
            return redirect(url_for('pacientes.detalle_paciente', id=id))
            
        Cita.crear(id, fecha, hora)
        flash('Nueva cita registrada exitosamente.', 'success')
        
    return redirect(url_for('pacientes.detalle_paciente', id=id))

@pacientes.route('/<int:id>/actualizar_cita/<int:cita_id>', methods=['POST'])
def actualizar_cita(id, cita_id):
    fecha = request.form.get('proxima_cita_fecha')
    hora = request.form.get('proxima_cita_hora')

    if not (9 <= int(hora.split(':')[0]) <= 19):
        flash('La hora debe estar entre las 9:00 AM y las 7:00 PM.', 'error')
        return redirect(url_for('pacientes.detalle_paciente', id=id))

    cita = Cita.query.get(cita_id)
    if cita:
        if datetime.now() > datetime.combine(cita.fecha, cita.hora):
            flash('No se puede actualizar una cita que ya ha pasado.', 'error')
            return redirect(url_for('pacientes.detalle_paciente', id=id))

    if not Cita.es_horario_disponible(fecha, hora, excluir_cita_id=cita_id):
        flash('El horario seleccionado ya no está disponible.', 'error')
        return redirect(url_for('pacientes.detalle_paciente', id=id))

    cita.fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
    cita.hora = datetime.strptime(hora, '%H:%M').time() if len(hora) == 5 else datetime.strptime(hora, '%H:%M:%S').time()
    db.session.commit()

    flash('Cita actualizada exitosamente.', 'success')
    return redirect(url_for('pacientes.detalle_paciente', id=id))

@pacientes.route('/disponibilidad_horas', methods=['GET'])
def disponibilidad_horas():
    fecha = request.args.get('fecha')
    citas_dia = Cita.query.filter_by(fecha=datetime.strptime(fecha, '%Y-%m-%d').date()).all()
    horas_ocupadas = [c.hora.strftime('%H:%M') for c in citas_dia if c.hora]
    return jsonify(horas_ocupadas)
